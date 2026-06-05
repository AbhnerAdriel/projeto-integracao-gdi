"""Etapa de extração dos arquivos XLSX.

As planilhas têm duas linhas iniciais de metadados: a linha 2 possui códigos das
perguntas e a linha 3 possui os nomes das colunas. Por isso, a leitura principal usa
`header=2`, considerando a terceira linha como cabeçalho.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd

from src.config import SOURCE_FILES
from src.utils import make_unique_column_names


def read_question_codes(path: Path) -> pd.DataFrame:
    """Lê os códigos das perguntas, mantendo a relação com os nomes originais."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    codes = list(next(worksheet.iter_rows(min_row=2, max_row=2, values_only=True)))
    original_columns = list(next(worksheet.iter_rows(min_row=3, max_row=3, values_only=True)))
    normalized_columns = make_unique_column_names(original_columns)
    return pd.DataFrame(
        {
            "codigo_planilha": codes,
            "campo_original": original_columns,
            "coluna_normalizada": normalized_columns,
        }
    )


def read_year_file(path: Path, year: int) -> pd.DataFrame:
    """Lê um arquivo anual da pesquisa e adiciona a coluna `ano_fonte`."""
    df = pd.read_excel(path, sheet_name=0, header=2)
    df.columns = make_unique_column_names(df.columns)
    df["ano_fonte"] = year
    df["arquivo_origem"] = path.name
    return df


def extract_all_sources(source_files: dict[int, Path] | None = None) -> pd.DataFrame:
    """Extrai e empilha os três anos de dados em um único DataFrame."""
    source_files = source_files or SOURCE_FILES
    frames: list[pd.DataFrame] = []
    for year, path in sorted(source_files.items()):
        if not path.exists():
            raise FileNotFoundError(f"Arquivo não encontrado: {path}")
        frames.append(read_year_file(path, year))
    return pd.concat(frames, ignore_index=True)
